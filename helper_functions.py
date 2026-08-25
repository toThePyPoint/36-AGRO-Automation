import pandas as pd
from openpyxl.styles import Font, PatternFill


def export_to_formatted_excel(
    df: pd.DataFrame, output_path, sheet_name: str = 'Sheet1'
) -> None:
    """Exports a pandas DataFrame to an Excel file with auto-fitted column widths

    and applied conditional formatting for quantity and comment columns.
    """
    # Define highlight styles (Excel default soft colors)
    red_fill = PatternFill(
        start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
    )
    red_font = Font(color='9C0006', bold=True)

    green_fill = PatternFill(
        start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'
    )
    green_font = Font(color='006100', bold=True)

    # 1. Export DataFrame to Excel using openpyxl
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]

        # 2. Auto-fit column widths
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(
                max_len + 3, 10
            )

        # 3. Highlight negative values in 'Ilość po wydaniu' column
        col_qty_name = 'Ilość po wydaniu'
        if col_qty_name in df.columns:
            col_idx = df.columns.get_loc(col_qty_name) + 1
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)) and cell.value < 0:
                    cell.fill = red_fill
                    cell.font = red_font

        # 4. Highlight non-empty comments other than '-' in 'komentarz' column
        col_comment_name = 'komentarz'
        if col_comment_name in df.columns:
            col_idx = df.columns.get_loc(col_comment_name) + 1
            for row_idx in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                val_str = (
                    str(cell.value).strip() if cell.value is not None else ''
                )
                if val_str and val_str != '-':
                    cell.fill = green_fill
                    cell.font = green_font